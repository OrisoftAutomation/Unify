import pytest
import time
import xlrd
import XLUtils
import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

@pytest.mark.usefixtures("setup")
class TestAppraisalOne:

    def test_appraisal(self):

        driver=self.driver
        responseStart = driver.execute_script("return window.performance.timing.responseStart")
        filepath="C:/Users/SRIN11081/Desktop/pytest load test_test/pytest load test_test/appraisal_test.xlsx" 
        wb_data = xlrd.open_workbook(filepath)
        wb_input = load_workbook(filepath)  #workbook
        wb_data = load_workbook(filepath)
        t_rows_input=XLUtils.getRowCountInput(filepath,'Inputs')
        t_row_element=XLUtils.getRowCountElement(filepath,'Data Elements')
        time_sheet = wb_data.get_sheet_by_name('Inputs')
        sheet_name_Inputs="Inputs"
        #print("Max Rows for Input :",t_rows_input)
        #print("Max Rows for Data Elements :",t_row_element)

        for r in range(2,t_rows_input, 1):
            #browser = WebDriver('chrome')
            #print("New R :", r)
            #inputs_col=3
            user_id = XLUtils.readData(filepath,sheet_name_Inputs, r, 4)
            password = XLUtils.readData(filepath,sheet_name_Inputs, r, 5)

            question_1 = XLUtils.readData(filepath,sheet_name_Inputs, r, 6)
            kpi_1 = XLUtils.readData(filepath,sheet_name_Inputs, r, 7)
            threshold_1 = XLUtils.readData(filepath,sheet_name_Inputs, r, 8)
            on_target_1 = XLUtils.readData(filepath,sheet_name_Inputs, r, 9)
            stretched_1 = XLUtils.readData(filepath,sheet_name_Inputs, r, 10)
            weightage_1 = XLUtils.readData(filepath,sheet_name_Inputs, r, 11)
            
            question_2 = XLUtils.readData(filepath,sheet_name_Inputs, r, 12)
            kpi_2 = XLUtils.readData(filepath,sheet_name_Inputs, r, 13)
            threshold_2 = XLUtils.readData(filepath,sheet_name_Inputs, r, 14)
            on_target_2 = XLUtils.readData(filepath,sheet_name_Inputs, r, 15)
            stretched_2 = XLUtils.readData(filepath,sheet_name_Inputs, r, 16)
            weightage_2 = XLUtils.readData(filepath,sheet_name_Inputs, r, 17)

            question_3 = XLUtils.readData(filepath,sheet_name_Inputs, r, 18)
            kpi_3 = XLUtils.readData(filepath,sheet_name_Inputs, r, 19)
            threshold_3 = XLUtils.readData(filepath,sheet_name_Inputs, r, 20)
            on_target_3 = XLUtils.readData(filepath,sheet_name_Inputs, r, 21)
            stretched_3 = XLUtils.readData(filepath,sheet_name_Inputs, r, 22)
            weightage_3 = XLUtils.readData(filepath,sheet_name_Inputs, r, 24)

            question_4 = XLUtils.readData(filepath,sheet_name_Inputs, r, 25)
            kpi_4 = XLUtils.readData(filepath,sheet_name_Inputs, r, 26)
            threshold_4 = XLUtils.readData(filepath,sheet_name_Inputs, r, 27)
            on_target_4 = XLUtils.readData(filepath,sheet_name_Inputs, r, 28)
            stretched_4 = XLUtils.readData(filepath,sheet_name_Inputs, r, 29)
            weightage_4 = XLUtils.readData(filepath,sheet_name_Inputs, r, 30)

            question_5 = XLUtils.readData(filepath,sheet_name_Inputs, r, 31)
            kpi_5 = XLUtils.readData(filepath,sheet_name_Inputs, r, 32)
            threshold_5 = XLUtils.readData(filepath,sheet_name_Inputs, r, 33)
            on_target_5 = XLUtils.readData(filepath,sheet_name_Inputs, r, 34)
            stretched_5 = XLUtils.readData(filepath,sheet_name_Inputs, r, 35)
            weightage_5 = XLUtils.readData(filepath,sheet_name_Inputs, r, 36)

            question_6 = XLUtils.readData(filepath,sheet_name_Inputs, r, 37)
            kpi_6 = XLUtils.readData(filepath,sheet_name_Inputs, r, 38)
            threshold_6 = XLUtils.readData(filepath,sheet_name_Inputs, r, 39)
            on_target_6 = XLUtils.readData(filepath,sheet_name_Inputs, r, 40)
            stretched_6 = XLUtils.readData(filepath,sheet_name_Inputs, r, 41)
            weightage_6 = XLUtils.readData(filepath,sheet_name_Inputs, r, 42)

            question_7 = XLUtils.readData(filepath,sheet_name_Inputs, r, 43)
            kpi_7 = XLUtils.readData(filepath,sheet_name_Inputs, r, 44)
            threshold_7 = XLUtils.readData(filepath,sheet_name_Inputs, r, 45)
            on_target_7 = XLUtils.readData(filepath,sheet_name_Inputs, r, 46)
            stretched_7 = XLUtils.readData(filepath,sheet_name_Inputs, r, 47)
            weightage_7 = XLUtils.readData(filepath,sheet_name_Inputs, r, 48)

            question_8 = XLUtils.readData(filepath,sheet_name_Inputs, r, 49)
            kpi_8 = XLUtils.readData(filepath,sheet_name_Inputs, r, 50)
            threshold_8 = XLUtils.readData(filepath,sheet_name_Inputs, r, 51)
            on_target_8 = XLUtils.readData(filepath,sheet_name_Inputs, r, 52)
            stretched_8 = XLUtils.readData(filepath,sheet_name_Inputs, r, 53)
            weightage_8 = XLUtils.readData(filepath,sheet_name_Inputs, r, 54)

            question_9 = XLUtils.readData(filepath,sheet_name_Inputs, r, 55)
            kpi_9 = XLUtils.readData(filepath,sheet_name_Inputs, r, 56)
            threshold_9 = XLUtils.readData(filepath,sheet_name_Inputs, r, 57)
            on_target_9 = XLUtils.readData(filepath,sheet_name_Inputs, r, 58)
            stretched_9 = XLUtils.readData(filepath,sheet_name_Inputs, r, 59)
            weightage_9 = XLUtils.readData(filepath,sheet_name_Inputs, r, 60)

            question_10 = XLUtils.readData(filepath,sheet_name_Inputs, r, 61)
            kpi_10 = XLUtils.readData(filepath,sheet_name_Inputs, r, 62)
            threshold_10 = XLUtils.readData(filepath,sheet_name_Inputs, r, 63)
            on_target_10 = XLUtils.readData(filepath,sheet_name_Inputs, r, 64)
            stretched_10 = XLUtils.readData(filepath,sheet_name_Inputs, r, 65)
            weightage_10 = XLUtils.readData(filepath,sheet_name_Inputs, r, 66)

            question_11 = XLUtils.readData(filepath,sheet_name_Inputs, r, 67)
            kpi_11 = XLUtils.readData(filepath,sheet_name_Inputs, r, 68)
            threshold_11 = XLUtils.readData(filepath,sheet_name_Inputs, r, 69)
            on_target_11 = XLUtils.readData(filepath,sheet_name_Inputs, r, 70)
            stretched_11 = XLUtils.readData(filepath,sheet_name_Inputs, r, 71)
            weightage_11 = XLUtils.readData(filepath,sheet_name_Inputs, r, 72)

            question_12 = XLUtils.readData(filepath,sheet_name_Inputs, r, 73)
            kpi_12 = XLUtils.readData(filepath,sheet_name_Inputs, r, 74)
            threshold_12 = XLUtils.readData(filepath,sheet_name_Inputs, r, 75)
            on_target_12 = XLUtils.readData(filepath,sheet_name_Inputs, r, 76)
            stretched_12 = XLUtils.readData(filepath,sheet_name_Inputs, r, 77)
            weightage_12 = XLUtils.readData(filepath,sheet_name_Inputs, r, 78)

#xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx

            #browser = WebDriver('chrome', reuse_browser=True)

            elem_userid=driver.find_element_by_name("ctl00$cntPlcHldrContent$txtUsrID")
            elem_userid.send_keys(user_id) #data_element_login7
            driver.implicitly_wait(20000) #time_sleep_login6
            print("User ID is entered ",user_id)

            elem_pwd=driver.find_element_by_name("ctl00$cntPlcHldrContent$txtUsrPwd") #data_element_login8
            elem_pwd.send_keys(password)
            #ActionChains(driver).move_to_element(elem_pwd).click(elem_pwd).send_keys(password).perform()
            #time.sleep(time_sleep[7].value) #time_sleep_login7
            print("Password is entered ",password)

            sign_in=driver.find_element_by_name("ctl00$cntPlcHldrContent$btnSignIn") ##data_element_login9
            sign_in.click()
            print("Sign In button is clicked")
            driver.implicitly_wait(20000)

            wait = WebDriverWait(driver, 2)
            open_menu = wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@id='header']/div[1]/div[2]/div[2]/img")))
            open_menu = driver.find_element_by_xpath("//*[@id='header']/div[1]/div[2]/div[2]/img")
            open_menu.click()

            driver.implicitly_wait(20000)

            appraisal_link=wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@id='header']/div[1]/div[4]/div[2]/div[2]/ul/li[3]/a")))
            appraisal_link=driver.find_element_by_xpath("//*[@id='header']/div[1]/div[4]/div[2]/div[2]/ul/li[3]/a")
            appraisal_link.click()
            driver.implicitly_wait(20000)

         

            appraisal_open_form = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#cntPlcHldrContent_cntPlcHldrContent_gvSearch > tbody > tr:nth-child(3) > td:nth-child(3) > a")))
            appraisal_open_form =driver.find_element_by_css_selector("#cntPlcHldrContent_cntPlcHldrContent_gvSearch > tbody > tr:nth-child(3) > td:nth-child(3) > a")
            ActionChains(driver).move_to_element(appraisal_open_form).click(appraisal_open_form).perform()
            driver.implicitly_wait(20000)
            
            start_time_open_name = time.time() 

            appraisal_open_name = wait.until(EC.visibility_of_element_located((By.XPATH, "//*[@id='tblAppraisal']/tbody/tr/td[2]/span/a")))
            appraisal_open_name =driver.find_element_by_xpath("//*[@id='tblAppraisal']/tbody/tr/td[2]/span/a")
            ActionChains(driver).move_to_element(appraisal_open_name).click(appraisal_open_name).perform()
            driver.implicitly_wait(20000)

            end_time_open_name = time.time()
            
            elapsed_time_open_name = end_time_open_name-start_time_open_name
            elapsed_time_open_name = time.strftime("%H:%M:%S", time.gmtime(elapsed_time_open_name))
            print(elapsed_time_open_name)
            time_output_open_name=time_sheet.cell(row=r,column=1).value=elapsed_time_open_name

#xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
            # question 1

            wait = WebDriverWait(driver, 2)
            appraisal_no_1 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_1 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_1).click(appraisal_no_1).send_keys(question_1).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_1 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_1 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_1).click(appraisal_kpi_1).send_keys(kpi_1).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_1 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_1 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_1).click(appraisal_threshold_1).send_keys(threshold_1).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_1 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_1 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_1).click(appraisal_on_target_1).send_keys(on_target_1).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_1 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_1 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_1).click(appraisal_stretched_1).send_keys(stretched_1).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage1 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage1 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage1).click(appraisal_weightage1).send_keys(weightage_1).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            
#xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
             #question 2-12

            wait = WebDriverWait(driver, 2)
            appraisal_no_2 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_2 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_2).click(appraisal_no_2).send_keys(question_2).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_2 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_2 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_2).click(appraisal_kpi_2).send_keys(kpi_2).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_2 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_2 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_2).click(appraisal_threshold_2).send_keys(threshold_2).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_2 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_2 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_2).click(appraisal_on_target_2).send_keys(on_target_2).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_2 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_2 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_2).click(appraisal_stretched_2).send_keys(stretched_2).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage2 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage2 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage2).click(appraisal_weightage2).send_keys(weightage_2).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_3 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_3 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_3).click(appraisal_no_3).send_keys(question_3).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_3 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_3 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_3).click(appraisal_kpi_3).send_keys(kpi_3).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_3 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_3 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_3).click(appraisal_threshold_3).send_keys(threshold_3).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_3 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_3 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_3).click(appraisal_on_target_3).send_keys(on_target_3).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_3 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_3 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_3).click(appraisal_stretched_3).send_keys(stretched_3).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage3 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage3 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage3).click(appraisal_weightage3).send_keys(weightage_3).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_4 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_4 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_4).click(appraisal_no_4).send_keys(question_4).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_4 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_4 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_4).click(appraisal_kpi_4).send_keys(kpi_4).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_4 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_4 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_4).click(appraisal_threshold_4).send_keys(threshold_4).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_4 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_4 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_4).click(appraisal_on_target_4).send_keys(on_target_4).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_4 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_4 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_4).click(appraisal_stretched_4).send_keys(stretched_4).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage4 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage4 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage4).click(appraisal_weightage4).send_keys(weightage_4).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_5 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_5 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_5).click(appraisal_no_5).send_keys(question_5).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_5 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_5 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_5).click(appraisal_kpi_5).send_keys(kpi_5).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_5 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_5 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_5).click(appraisal_threshold_5).send_keys(threshold_5).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_5 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_5 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_5).click(appraisal_on_target_5).send_keys(on_target_5).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_5 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_5 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_5).click(appraisal_stretched_5).send_keys(stretched_5).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage5 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage5 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage5).click(appraisal_weightage5).send_keys(weightage_5).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_6 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_6 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_6).click(appraisal_no_6).send_keys(question_6).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_6 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_6 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_6).click(appraisal_kpi_6).send_keys(kpi_6).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_6 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_6 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_6).click(appraisal_threshold_6).send_keys(threshold_6).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_6 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_6 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_6).click(appraisal_on_target_6).send_keys(on_target_6).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_6 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_6 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_6).click(appraisal_stretched_6).send_keys(stretched_6).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage6 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage6 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage6).click(appraisal_weightage6).send_keys(weightage_6).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_7 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_7 =driver.find_element_by_id("txtQuestionDesc_0")  
            ActionChains(driver).move_to_element(appraisal_no_7).click(appraisal_no_7).send_keys(question_7).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_7 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_7 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_7).click(appraisal_kpi_7).send_keys(kpi_7).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_7 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_7 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_7).click(appraisal_threshold_7).send_keys(threshold_7).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_7 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_7 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_7).click(appraisal_on_target_7).send_keys(on_target_7).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_7 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_7 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_7).click(appraisal_stretched_7).send_keys(stretched_7).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage7 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage7 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage7).click(appraisal_weightage7).send_keys(weightage_7).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_8 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_8 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_8).click(appraisal_no_8).send_keys(question_8).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_8 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_8 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_8).click(appraisal_kpi_8).send_keys(kpi_8).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_8 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_8 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_8).click(appraisal_threshold_8).send_keys(threshold_8).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_8 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_8 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_8).click(appraisal_on_target_8).send_keys(on_target_8).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_8 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_8 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_8).click(appraisal_stretched_8).send_keys(stretched_8).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage8 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage8 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage8).click(appraisal_weightage8).send_keys(weightage_8).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_9 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_9 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_9).click(appraisal_no_9).send_keys(question_9).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_9 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_9 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_9).click(appraisal_kpi_9).send_keys(kpi_9).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_9 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_9 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_9).click(appraisal_threshold_9).send_keys(threshold_9).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_9 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_9 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_9).click(appraisal_on_target_9).send_keys(on_target_9).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_9 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_9 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_9).click(appraisal_stretched_9).send_keys(stretched_9).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage9 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage9 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage9).click(appraisal_weightage9).send_keys(weightage_9).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_10 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_10 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_10).click(appraisal_no_10).send_keys(question_10).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_10 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_10 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_10).click(appraisal_kpi_10).send_keys(kpi_10).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_10 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_10 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_10).click(appraisal_threshold_10).send_keys(threshold_10).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_10 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_10 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_10).click(appraisal_on_target_10).send_keys(on_target_10).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_10 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_10 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_10).click(appraisal_stretched_10).send_keys(stretched_10).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage10 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage10 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage10).click(appraisal_weightage10).send_keys(weightage_10).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_11 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_11 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_11).click(appraisal_no_11).send_keys(question_11).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_11 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_11 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_11).click(appraisal_kpi_11).send_keys(kpi_11).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_11 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_11 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_11).click(appraisal_threshold_11).send_keys(threshold_11).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_11 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_11 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_11).click(appraisal_on_target_11).send_keys(on_target_11).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_11 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_11 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_11).click(appraisal_stretched_11).send_keys(stretched_11).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage11 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage11 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage11).click(appraisal_weightage11).send_keys(weightage_11).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()


            wait = WebDriverWait(driver, 2)
            appraisal_no_12 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_no_12 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_no_12).click(appraisal_no_12).send_keys(question_12).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            appraisal_kpi_12 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_kpi_12 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_kpi_12).click(appraisal_kpi_12).send_keys(kpi_12).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_threshold_12 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_threshold_12 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_threshold_12).click(appraisal_threshold_12).send_keys(threshold_12).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_on_target_12 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_on_target_12 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_on_target_12).click(appraisal_on_target_12).send_keys(on_target_12).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_stretched_12 = wait.until(EC.visibility_of_element_located((By.ID, "txtQuestionDesc_0")))
            appraisal_stretched_12 =driver.find_element_by_id("txtQuestionDesc_0")
            ActionChains(driver).move_to_element(appraisal_stretched_12).click(appraisal_stretched_12).send_keys(stretched_12).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()

            appraisal_weightage12 = wait.until(EC.visibility_of_element_located((By.NAME, "txtWeightMark_0")))
            appraisal_weightage12 =driver.find_element_by_name("txtWeightMark_0")
            ActionChains(driver).move_to_element(appraisal_weightage12).click(appraisal_weightage12).send_keys(weightage_12).perform()
            driver.implicitly_wait(20000)

            elemClickSomewhere = driver.find_element_by_xpath("//*[@id='tblGeneral']/thead/tr[1]/td[2]")
            elemClickSomewhere.click()
            
            #question 2-12
#xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
            elemHome=driver.find_element_by_tag_name('body').send_keys(Keys.HOME)
            driver.implicitly_wait(20000)
            
#xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx

            elemOk = wait.until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/div[3]/div/button/span")))
            wait2 = WebDriverWait(driver, 3)
            elemSave = wait2.until(EC.visibility_of_element_located((By.NAME, "ctl00$ctl00$ctl00$cntPlcHldrContent$btnSave")))
            start_save_time = time.time()
            ActionChains(driver).move_to_element(elemSave).click().perform()
            driver.implicitly_wait(20000)

            elemOk = driver.find_element_by_xpath("/html/body/div[2]/div[3]/div/button/span")
            end_save_time = time.time()
            elapsed_save_time = end_save_time-start_save_time
            elapsed_save_time = time.strftime("%H:%M:%S", time.gmtime(elapsed_save_time))
            print(elapsed_save_time)
            time_output_save=time_sheet.cell(row=r,column=2).value=elapsed_save_time
            elemOk.click()
            

            

            '''wait3 = WebDriverWait(driver, 3)
            elemSignOut = wait3.until(EC.visibility_of_element_located((By.XPATH, data_elements[20].value)))
            elemSignOut = driver.find_element_by_xpath(data_elements[20].value)
            elemSignOut.click()
            time.sleep(0.5)'''
            #end_time = time.time()
            #elapsed_time = end_time - start_time


            
            #print ("ELAPSED TIME: ", elapsed_time)
            #elapsed_time = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))

            #print("Elapsed time:", elapsed_time)

            
            #time_output=time_sheet.cell(row=r,column=2).value=elapsed_time
   
            
            #time.sleep(3)
            
            #appraisal_form1=

            '''
            open_form = wait.until(EC.element_to_be_clickable((By.XPATH, data_elements[6].value)))
            open_form = driver.find_element_by_xpath(data_elements[6].value)
            open_form.click()
            time.sleep(2)'''

            domComplete = driver.execute_script("return window.performance.timing.domComplete")

            frontendPerformance = domComplete - responseStart

            #frontendPerformance=time.strftime("%S", time.gmtime(frontendPerformance))
            print (frontendPerformance)

            time_output_total=time_sheet.cell(row=r,column=3).value=frontendPerformance
            wb_data.save(filepath)

            time.sleep(2)

